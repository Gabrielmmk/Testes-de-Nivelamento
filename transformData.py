import pandas as pd
import tabula
import openpyxl
import zipfile
import os
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Adicione seu nome aqui
SEU_NOME = "Gabriel"  # Substitua pelo seu nome


def pdf_para_dataframe(rota_arquivo, paginas="3-181"):
    """
    Extrai tabelas de PDF e retorna um DataFrame consolidado.
    """
    try:
        print(f"Extraindo tabelas do PDF: {rota_arquivo}...")
        tabelas = tabula.read_pdf(
            rota_arquivo,
            pages=paginas,
            multiple_tables=True,
            guess=False,
            silent=True,
            lattice=True
        )

        if not tabelas:
            raise ValueError("Nenhuma tabela encontrada no PDF.")

        dataframe_combinado = pd.concat(tabelas, ignore_index=True)

        if len(dataframe_combinado.columns) == 13:
            colunas_corretas = [
                'PROCEDIMENTO', 'RN', 'VIGÊNCIA', 'ODONTOLÓGICA', 'AMBULATORIAL',
                'HCO', 'HSO', 'REF', 'PAC', 'DUT', 'SUBGRUPO', 'GRUPO', 'CAPÍTULO'
            ]
            dataframe_combinado.columns = colunas_corretas
        else:
            print(
                f"Aviso: Número de colunas inesperado ({len(dataframe_combinado.columns)}). Mantendo nomes originais.")

        return dataframe_combinado

    except Exception as e:
        print(f"Erro ao extrair tabelas do PDF: {e}")
        raise


def aplicar_formatacao_excel(nome_arquivo):
    """
    Aplica formatação profissional ao arquivo Excel.
    """
    try:
        wb = openpyxl.load_workbook(nome_arquivo)
        sheet = wb.active

        azul_claro = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
        branco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cabecalho_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        cabecalho_font = openpyxl.styles.Font(color="FFFFFF", bold=True)

        borda = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for cell in sheet[1]:
            cell.fill = cabecalho_fill
            cell.font = cabecalho_font
            cell.border = borda
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = borda
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.fill = azul_claro if row_idx % 2 == 0 else branco

        for col in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        sheet.freeze_panes = "A2"
        wb.save(nome_arquivo)
        print(f"Formatação aplicada com sucesso em {nome_arquivo}")

    except Exception as e:
        print(f"Erro ao aplicar formatação: {e}")


def salvar_como_excel_formatado(dataframe, nome_arquivo="tabela_formatada.xlsx"):
    """
    Salva o DataFrame como Excel com formatação profissional.
    """
    try:
        dataframe.to_excel(nome_arquivo, index=False, engine='openpyxl')
        aplicar_formatacao_excel(nome_arquivo)
    except Exception as e:
        print(f"Erro ao salvar arquivo Excel: {e}")


def salvar_como_csv(dataframe, nome_arquivo="tabela_consolidada.csv"):
    """
    Salva o DataFrame como CSV com encoding UTF-8.
    """
    try:
        dataframe.to_csv(nome_arquivo, sep=";", index=False, encoding="utf-8-sig")
        print(f"CSV salvo com sucesso: {nome_arquivo}")
        return nome_arquivo
    except Exception as e:
        print(f"Erro ao salvar CSV: {e}")
        return None


def compactar_csv_para_zip(csv_path, nome_zip=None):
    """
    Compacta o arquivo CSV em um ZIP com o nome especificado.
    """
    if nome_zip is None:
        nome_zip = f"Teste_{SEU_NOME}.zip"

    try:
        with zipfile.ZipFile(nome_zip, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(csv_path, os.path.basename(csv_path))
        print(f"Arquivo CSV compactado em: {nome_zip}")
        return nome_zip
    except Exception as e:
        print(f"Erro ao compactar arquivo: {e}")
        return None


if __name__ == "__main__":
    try:
        # Configurações
        pdf_path = 'C:/Users/gabri/PycharmProjects/testeDeNivelamento/downloads/Anexo_I_Rol_2021RN_465.2021_RN627L.2024.pdf'
        output_excel = "ProcedimentoExcel.xlsx"
        output_csv = "ProcedimentoCSV.csv"
        output_zip = f"Teste_{SEU_NOME}.zip"

        # Processamento
        df = pdf_para_dataframe(pdf_path)

        # Exportação
        csv_file = salvar_como_csv(df, output_csv)
        if csv_file:
            compactar_csv_para_zip(csv_file, output_zip)

        salvar_como_excel_formatado(df, output_excel)

    except Exception as e:
        print(f"Erro no processamento: {e}")